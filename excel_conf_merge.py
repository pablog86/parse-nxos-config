import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

absolute = False #Compare static routes by absolute Network + Next-hop

pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', None)
pd.set_option('display.max_colwidth', -1)

#----------------------------------------------------
def excel_merge (writer, xls1, xls2, sheet, index=None, sort=None, compares=None):
	df_xls1 = pd.read_excel(xls1, sheet).fillna("#-")
	df_xls2 = pd.read_excel(xls2, sheet).fillna("#-")

	if index != None and (len(df_xls1.index)!=0 and len(df_xls2.index)!=0):
		if True in list(df_xls1.duplicated([index])) or True in list(df_xls2.duplicated([index])):
			if sheet == "Statics":
				if absolute:
					df_xls1["new_index"] = df_xls1[index] + df_xls1["Next-hop"]
					df_xls2["new_index"] = df_xls2[index] + df_xls2["Next-hop"]
					index = "new_index"
				else:
					l1 = list(df_xls1.duplicated([index]))
					l1 = [i for i,x in enumerate(list(l1)) if x==True]
					for i in l1:
						df_xls1[index][i] = df_xls1[index][i] + " " #add space to change de string a break the dupplication of index
					l2 = list(df_xls2.duplicated([index]))
					l2 = [i for i,x in enumerate(list(l2)) if x==True]
					for i in l2:
						df_xls2[index][i] = df_xls2[index][i] + " " #add space to change de string a break the dupplication of index
			if sheet == "IP_ACLs":
				df_xls1["new_index"] = df_xls1[index].astype(str)+df_xls1["src"]+df_xls1["src_port"].astype(str)+df_xls1["dst"]+df_xls1["dst_port"].astype(str)+df_xls1["action"]+df_xls1["protocol"]#+df_xls1["rule"].astype(str)
				df_xls2["new_index"] = df_xls2[index].astype(str)+df_xls2["src"]+df_xls2["src_port"].astype(str)+df_xls2["dst"]+df_xls2["dst_port"].astype(str)+df_xls2["action"]+df_xls2["protocol"]#+df_xls2["rule"].astype(str)
				index = "new_index"
				if df_xls1.new_index.duplicated().any():
					print("Duplicated in A: ", df_xls1.new_index.duplicated())
					df_xls1 = df_xls1.drop_duplicates(subset=['new_index'])
				if df_xls2.new_index.duplicated().any():
					print("Duplicated in B: ", df_xls2.new_index.duplicated())
					df_xls2 = df_xls2.drop_duplicates(subset=['new_index'])

		df_xls1 = df_xls1.set_index(index)
		df_xls2 = df_xls2.set_index(index)

	df = pd.concat([df_xls1, df_xls2], axis=1).fillna("X")

	if sort != None:
		df = df.sort_values(by=sort, ascending=True)
	if compares != None:
		col = []
		for compare in compares:
			col.append([i for i,x in enumerate(list(df.columns)) if x==compare])
		if len(col) == 1:
			df['diff'] = np.where(df.iloc[:,col[0][0]] == df.iloc[:,col[0][1]], 'True', 'False') 
		else:
			df['diff'] = np.where(np.logical_or(df.iloc[:,col[0][0]] == df.iloc[:,col[0][1]], df.iloc[:,col[1][0]] == df.iloc[:,col[1][1]]), 'True', 'False')
	
	df = df.replace(to_replace="#-", value=np.nan)

	#if sheet == 'Statics':   #Test
	#	comparing(df, 'Next-hop')

	df.to_excel(writer, sheet_name=sheet, startrow = 1)
	#headers(writer.sheets[sheet], sheet)
	formatting(writer.sheets[sheet], sheet)
	headers(writer.sheets[sheet], sheet)
	#writer.save()
#----------------------------------------------------
def headers (ws, sheet):
	ws["B1"] = "Nexus A"
	font = Font(name = "Arial", size = 10, bold = True, color='ffffff')
	alignment = Alignment(horizontal="center", vertical="center")
	thin = Side(border_style = "thin", color = "000000")
	border = Border(bottom = thin, top = thin, right = thin, left = thin)
	color = PatternFill(start_color = "365F91", end_color = "365F91", fill_type = "solid")
	ws["B1"].font = font
	ws["B1"].border = border
	ws["B1"].fill = color
	ws["B1"].alignment = alignment
	if sheet == "VLANs":	
		ws["C1"] = "Nexus B"
		ws["C1"].font = font
		ws["C1"].border = border
		ws["C1"].fill = color
		ws["C1"].alignment = alignment
	if sheet == "SVIs":
		ws["G1"] = "Nexus B"
		ws["G1"].font = font
		ws["G1"].border = border
		ws["G1"].fill = color
		ws["G1"].alignment = alignment
		ws.merge_cells(start_row = 1, start_column = 2, end_row = 1, end_column = 6)
		ws.merge_cells(start_row = 1, start_column = 7, end_row = 1, end_column = 11)
	if sheet == "Ints":
		ws["I1"] = "Nexus B"
		ws["I1"].font = font
		ws["I1"].border = border
		ws["I1"].fill = color
		ws["I1"].alignment = alignment
		ws.merge_cells(start_row = 1, start_column = 2, end_row = 1, end_column = 8)
		ws.merge_cells(start_row = 1, start_column = 9, end_row = 1, end_column = 15)
	if sheet == "Po":
		ws["H1"] = "Nexus B"
		ws["H1"].font = font
		ws["H1"].border = border
		ws["H1"].fill = color
		ws["H1"].alignment = alignment
		ws.merge_cells(start_row = 1, start_column = 2, end_row = 1, end_column = 7)
		ws.merge_cells(start_row = 1, start_column = 8, end_row = 1, end_column = 15)
	if sheet == "Statics":
		ws["J1"] = "Nexus B"
		ws["J1"].font = font
		ws["J1"].border = border
		ws["J1"].fill = color
		ws["J1"].alignment = alignment
		ws.merge_cells(start_row = 1, start_column = 2, end_row = 1, end_column = 9)
		ws.merge_cells(start_row = 1, start_column = 10, end_row = 1, end_column = 15)
	if sheet == "IP_ACLs":
		ws["O1"] = "Nexus B"
		ws["O1"].font = font
		ws["O1"].border = border
		ws["O1"].fill = color
		ws["O1"].alignment = alignment
		ws.merge_cells(start_row = 1, start_column = 2, end_row = 1, end_column = 14)
		ws.merge_cells(start_row = 1, start_column = 15, end_row = 1, end_column = 26)
#----------------------------------------------------
def formatting (ws, sheet):
	i = 0
	cmax = [6] * ws.max_column
	max_column = cmax
	thin = Side(border_style = "thin", color = "000000")
	color = PatternFill(start_color = "365F91", end_color = "365F91", fill_type = "solid")
	font = Font(name = "Arial", size = 10, bold = True, color='ffffff')
	for r in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=2):
		for c in r:
			c.border = Border(bottom = thin, top = thin, right = thin, left = thin)
			c.font = font
			c.fill = color
	font = Font(name = "Arial", size = 10)
	for r in ws.iter_rows(min_row=3, max_col=ws.max_column, max_row=ws.max_row):
		for c in r:
			c.border = Border(bottom = thin, top = thin, right = thin, left = thin)
			c.font = font
			if c.value == "X":
				c.fill = PatternFill(start_color = "FB2C57", end_color = "FB2C57", fill_type = "solid")
#Ajuste del ancho de columnas
			try:
				lenght = len(c.value)
			except TypeError as uni:
				lenght = 0
			cmax[i] = max(lenght, cmax[i])
			i += 1
		i = 0
		for j in range(ws.max_column):
			max_column[j] = max(max_column[j], cmax[j]) 
	for j in range(len(max_column)):
		ws.column_dimensions[get_column_letter(j+1)].width = min(max_column[j], 55)
#----------------------------------------------------
#def comparing (df, column):	#Test
#	a = [i for i,x in enumerate(list(df.columns)) if x==column]
#	for i in range(0,len(df.index)):
#		if df.iloc[i,a[0]] != df.iloc[i,a[1]]:
#			print(i)
#----------------------------------------------------
def write_excel (path1, path2):
	wb = openpyxl.Workbook()
	xls1 = pd.ExcelFile(path1)
	xls2 = pd.ExcelFile(path2)
	wbname1 = path1.split("/")[-1] 
	wbname1 = wbname1[:wbname1.rfind(".")]
	wbname2 = path2.split("/")[-1] 
	wbname2 = wbname2[:wbname2.rfind(".")]
	wbname = wbname1 + " - " + wbname2 + ".xlsx"
	print("Nexus A: ", wbname1)
	print("Nexus B: ", wbname2)
	with pd.ExcelWriter(wbname, engine="openpyxl") as writer:
		#def excel_merge (writer, xls1, xls2, sheet, index=None, sort=None, compares=None):
		excel_merge (writer, xls1, xls2, "VLANs", "VLAN", "VLAN", ["NAME"])
		excel_merge (writer, xls1, xls2, "SVIs", "SVI", "SVI", ["VIP"])
		excel_merge (writer, xls1, xls2, "Ints", "Interface", "Interface", ["VLANs/IP"])
		excel_merge (writer, xls1, xls2, "Po", "Interface", "Interface", ["VLANs/IP", "VPC"])
		#try:
		excel_merge (writer, xls1, xls2, "Statics", "Network", compares=["Next-hop"])
		#except:
		#	print("Multiples statics to same Network")
		#	excel_merge (writer, xls1, xls2, "Statics")
		excel_merge (writer, xls1, xls2, "IP_ACLs", "Name")
	print("Generando el archivo: ", wbname)
#----------------------------------------------------

if __name__ == "__main__":
	import tkinter as tk
	from tkinter import filedialog
	root = tk.Tk()
	root.withdraw()
	path1 = filedialog.askopenfilename()
	path2 = filedialog.askopenfilename()
	write_excel (path1, path2)




