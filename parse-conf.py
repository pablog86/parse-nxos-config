from ciscoconfparse import CiscoConfParse
from pathlib import Path
import re
import tkinter as tk
from tkinter import filedialog

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

import os
from os import listdir
from os.path import isfile, join

# ----------------------------------------------------


def formatting(ws, sheet):
    i = 0
    cmax = [6] * ws.max_column
    max_column = cmax
    thin = Side(border_style="thin", color="000000")
    for r in ws["A1:"+chr(ws.max_column+64)+str(ws.max_row)]:
        for c in r:
            c.border = Border(bottom=thin, top=thin, right=thin, left=thin)
            if c.value == "X":
                c.fill = PatternFill(start_color="FB2C57",
                                     end_color="FB2C57", fill_type="solid")
# Ajuste del ancho de columnas
            try:
                lenght = len(c.value)
            except TypeError as uni:
                lenght = 0
            cmax[i] = max(lenght, cmax[i])
            i += 1
        i = 0
        for j in range(ws.max_column):
            max_column[j] = max(max_column[j], cmax[j])
    for j in range(len(max_column)):
        ws.column_dimensions[chr(j+65)].width = max_column[j]+2
# ----------------------------------------------------
# VLANs information page


def vlan_sheet():
    row1 = 0
    row2 = 1
    df_vlan = pd.DataFrame(columns=['VLAN', 'NAME'])
    df_vlan_aux = pd.DataFrame(columns=['VLAN', 'NAME'])
    for p_obj in parse.find_objects('^vlan'):
        if p_obj.text[:18] != "vlan configuration":
            row1 += 1
            if len(p_obj.text[5:]) > 4:
                vlan_list = p_obj.text[5:].split(",")
                for v in vlan_list:
                    if "-" in v:
                        for j in range(int(v.split("-")[0]), int(v.split("-")[1])+1):
                            df_vlan_aux.loc[row2] = [int(j), np.nan]
                            row2 += 1
                    else:
                        df_vlan_aux.loc[row2] = [int(v), np.nan]
                        row2 += 1
            else:
                for c_obj in p_obj.children:
                    df_vlan.loc[row1] = [int(p_obj.text[5:]), str(c_obj)[str(
                        c_obj).find("name") + 5:str(c_obj).find("' (parent")]]
    df_vlan = df_vlan.set_index("VLAN")
    df_vlan_aux = df_vlan_aux.set_index("VLAN")
    if not df_vlan.index.equals(df_vlan_aux.index):
        for i in df_vlan_aux.index:
            df_vlan = df_vlan.combine_first(df_vlan_aux)
    return df_vlan
# ----------------------------------------------------
# VLANs Interfaces information page


def svi_sheet():
    row1 = 0
    df_svi = pd.DataFrame(
        columns=['SVI', 'Description', "IP", "VIP", "VRF", "ACL"])
    for p_obj in parse.find_objects('^interface Vlan')[1:]:
        desc, ipa, ipv, vrf, ipacl = np.nan, np.nan, np.nan, np.nan, np.nan
        row1 += 1
        for c_obj in p_obj.children:
            if "description" in str(c_obj):
                desc = str(c_obj)[str(c_obj).find(
                    "description") + 12:str(c_obj).find("' (parent")]
            if "ip address" in str(c_obj):
                ipa = str(c_obj)[str(c_obj).find(
                    "ip address") + 11:str(c_obj).find("' (parent")]
            if "hsrp" in str(c_obj):
                for h_obj in c_obj.children:
                    if "ip" in str(h_obj):
                        ipv = str(h_obj)[str(h_obj).find(
                            "ip") + 3:str(h_obj).find("' (parent")]
            if "vrf member" in str(c_obj):
                vrf = str(c_obj)[str(c_obj).find(
                    "vrf member") + 11:str(c_obj).find("' (parent")]
            if "ip access-group" in c_obj.text:
                ipacl = c_obj.text[18:].split(
                    " ")[0] + " (" + c_obj.text[18:].split(" ")[1] + ")"
        df_svi.loc[row1] = [int(str(p_obj)[str(p_obj).find(
            "interface Vlan") + 14:-2]), desc, ipa, ipv, vrf, ipacl]
    df_svi = df_svi.set_index("SVI")
    return df_svi
# ----------------------------------------------------
# Interfaces information page


def int_sheet(int_type, df_int_prev=pd.DataFrame(columns=['Interface', 'Description', "Type", "VLANs/IP", "Po", "Status", "VRF", "ACL"])):
    row1 = 0
    df_int = df_int_prev
    if df_int_prev.empty == False:
        row1 = df_int_prev.shape[0]
    for p_obj in parse.find_objects(int_type)[1:]:
        desc, typ, vlanip, po, status, vrf, ipacl = np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan
        row1 += 1
        for c_obj in p_obj.children:
            if "description" in str(c_obj):
                desc = str(c_obj)[str(c_obj).find(
                    "description") + 12:str(c_obj).find("' (parent")]
            try:
                if "access vlan" in str(c_obj):
                    typ = "switchport"
                    vlanip = str(c_obj)[str(c_obj).find(
                        "access vlan") + 12:str(c_obj).find("' (parent")]
                if "trunk allowed vlan" in str(c_obj):
                    if "add" in str(c_obj):
                        # Extend info
                        vlanip += ","+str(c_obj)[
                            str(c_obj).find("trunk allowed vlan add") +
                            23:str(c_obj).find("' (parent")]
                    else:
                        typ = "switchport"
                        vlanip = str(c_obj)[str(c_obj).find(
                            "trunk allowed vlan") + 19:str(c_obj).find("' (parent")]
                if "ip address" in str(c_obj):
                    vlanip = str(c_obj)[str(c_obj).find(
                        "ip address") + 11:str(c_obj).find("' (parent")]
            except Exception as ex:
                print("Problem with line ", str(p_obj), ex)
                # TODO: add something to the table
            if "channel-group" in str(c_obj):
                po = str(c_obj)[str(c_obj).find(
                    "channel-group") + 14:str(c_obj).find("' (parent")]
            if "no shutdown" in str(c_obj):
                status = "no shutdown"
            if "'  shutdown' (" in str(c_obj):
                status = "no shutdown"
            if "vrf member" in str(c_obj):
                vrf = str(c_obj)[str(c_obj).find(
                    "vrf member") + 11:str(c_obj).find("' (parent")]
            if "ip access-group" in c_obj.text:
                ipacl = c_obj.text[18:].split(
                    " ")[0] + " (" + c_obj.text[18:].split(" ")[1] + ")"
        df_int.loc[row1] = [str(p_obj)[str(p_obj).find(
            "interface") + 10:-2], desc, typ, vlanip, po, status, vrf, ipacl]
    # df_int = df_int.set_index("Interface") #keep the table size the same
    return df_int
# ----------------------------------------------------
# Port-channels information page


def po_sheet(int_type):
    row1 = 0
    df_po = pd.DataFrame(columns=[
                         'Interface', 'Description', "Type", "VLANs/IP", "Status", "VRF", "ACL", "VPC"])
    for p_obj in parse.find_objects(int_type):
        desc, typ, vlanip, status, vrf, ipacl, vpc = np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan
        row1 += 1
        for c_obj in p_obj.children:
            if "description" in str(c_obj):
                desc = str(c_obj)[str(c_obj).find(
                    "description") + 12:str(c_obj).find("' (parent")]
            try:
                if "access vlan" in str(c_obj):
                    typ = "switchport"
                    vlanip = str(c_obj)[str(c_obj).find(
                        "access vlan") + 12:str(c_obj).find("' (parent")]
                if "trunk allowed vlan" in str(c_obj):
                    if "add" in str(c_obj):
                        # Extend info
                        vlanip += ","+str(c_obj)[
                            str(c_obj).find("trunk allowed vlan add") +
                            23:str(c_obj).find("' (parent")]
                    else:
                        typ = "switchport"
                        vlanip = str(c_obj)[str(c_obj).find(
                            "trunk allowed vlan") + 19:str(c_obj).find("' (parent")]
                if "ip address" in str(c_obj):
                    typ = "routed"
                    vlanip = str(c_obj)[str(c_obj).find(
                        "ip address") + 11:str(c_obj).find("' (parent")]
            except Exception as ex:
                print("Problem with line ", str(p_obj), ex)
                # TODO: add something to the table
            if "no shutdown" in str(c_obj):
                status = "no shutdown"
            if "'  shutdown' (" in str(c_obj):
                status = "shutdown"
            if "vrf member" in str(c_obj):
                vrf = str(c_obj)[str(c_obj).find(
                    "vrf member") + 11:str(c_obj).find("' (parent")]
            if "vpc" in str(c_obj):
                vpc = c_obj.text[6:]
            if "ip access-group" in c_obj.text:
                ipacl = c_obj.text[18:].split(
                    " ")[0] + " (" + c_obj.text[18:].split(" ")[1] + ")"
        df_po.loc[row1] = [str(p_obj)[str(p_obj).find(
            "interface") + 10:-2], desc, typ, vlanip, status, vrf, ipacl, vpc]
    df_po = df_po.set_index("Interface")
    return df_po
# ----------------------------------------------------
# Static Routing


def static_sheet():
    row1 = 0
    df_route = pd.DataFrame(
        columns=['Network', 'Next-hop', "pref", "name", "tag", "track", "VRF"])
    for p_obj in parse.find_objects('^ip route'):
        net, nh, pref, name, tag, track, vrf = np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan
        row1 += 1
        route = p_obj.text[9:]
        route = route.split(" ")
        net = route.pop(0)
        nh = route.pop(0)
        if "name" in route:
            name = route.pop(route.index("name")+1)
            route.pop(route.index("name"))
        if "tag" in route:
            tag = route.pop(route.index("tag")+1)
            route.pop(route.index("tag"))
        if "track" in route:
            track = route.pop(route.index("track")+1)
            route.pop(route.index("track"))
        if len(route) != 0:
            pref = route[0]
        df_route.loc[row1] = [net, nh, pref, name, tag, track, vrf]
    for p_obj in parse.find_objects('^vrf context'):
        vrf = p_obj.text
        for c_obj in p_obj.children:
            net, nh, pref, name, tag, track = np.nan, np.nan, np.nan, np.nan, np.nan, np.nan
            row1 += 1
            route = c_obj.text[11:]
            route = route.split(" ")
            net = route.pop(0)
            nh = route.pop(0)
            if "name" in route:
                name = route.pop(route.index("name")+1)
                route.pop(route.index("name"))
            if "tag" in route:
                tag = route.pop(route.index("tag")+1)
                route.pop(route.index("tag"))
            if "track" in route:
                track = route.pop(route.index("track")+1)
                route.pop(route.index("track"))
            if len(route) != 0:
                pref = route[0]
            df_route.loc[row1] = [net, nh, pref, name, tag, track, vrf]
    return df_route
# ----------------------------------------------------
# IP ACLs


def ipacl_sheet():
    row1 = 0
    df_ipacl = pd.DataFrame(columns=['Name', "rule", "action", "protocol", "src",
                            "src_oper", "src_port", "dst", "dst_oper", "dst_port", "flags", "statistics"])
    for p_obj in parse.find_objects('^ip access-list '):
        name, rule, action, protocol, src, srco, srcp, dst, dsto, dstp, flag, stats = np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan
        row1 += 1
        name = p_obj.text[15:]
        for c_obj in p_obj.children:
            if "remark" in c_obj.text:
                acl = c_obj.text[2:].split(" ")
                rule = acl[0]
                action = acl[1]
                name = acl[2]
            else:
                try:
                    acl = c_obj.text[2:].split(" ")
                    if acl[0] == "statistics":
                        stats = acl[1]
                    else:
                        rule = acl.pop(0)
                        action = acl.pop(0)
                        protocol = acl.pop(0)
                        if "addrgroup" in acl[0]:
                            acl.pop(0)
                            src = acl.pop(0)
                        else:
                            src = acl.pop(0)
                        if "eq" in acl[0] or "gt" in acl[0] or "lt" in acl[0] or "neq" in acl[0] or "range" in acl[0]:
                            srco = acl.pop(0)
                            srcp = acl.pop(0)
                            if srco == "range":
                                srcp = srcp+"-"+acl.pop(0)
                        if "addrgroup" in acl[0]:
                            acl.pop(0)
                            dst = acl.pop(0)
                        else:
                            dst = acl.pop(0)
                        if "eq" in acl[0] or "gt" in acl[0] or "lt" in acl[0] or "neq" in acl[0] or "range" in acl[0]:
                            dsto = acl.pop(0)
                            dstp = acl.pop(0)
                            if dsto == "range":
                                dstp = dstp+"-"+acl.pop(0)
                        for i in range(len(acl)):
                            if flag == np.nan:
                                flag = flag + acl[i]
                            else:
                                flag = acl[i]
                        df_ipacl.loc[row1] = [name, rule, action, protocol,
                                              src, srco, srcp, dst, dsto, dstp, flag, stats]
                        row1 += 1
                except IndexError as uni:
                    # Maybe standar acl from catalyst?
                    src = protocol
                    protocol = np.nan
    return df_ipacl
# ----------------------------------------------------


def conf_excel(path, mode):
    df_vlan = vlan_sheet()
    df_svi = svi_sheet()
    #
    if mode == "nxos":
        df_int = int_sheet('^interface Ethernet')
    elif mode == "ios":
        df_int = int_sheet('^interface GigabitEthernet')
        df_int = int_sheet('^interface TenGigabitEthernet', df_int)
    df_int = df_int.set_index("Interface")
    #
    if mode == "nxos":
        df_po = po_sheet('^interface port-channel')
    elif mode == "ios":
        df_po = po_sheet('^interface Port-channel')
    df_route = static_sheet()
    df_ipacl = ipacl_sheet()
    wb = openpyxl.Workbook()
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_vlan.to_excel(writer, sheet_name="VLANs", startrow=0)
        formatting(writer.sheets["VLANs"], "VLANs")
        df_svi.to_excel(writer, sheet_name="SVIs", startrow=0)
        formatting(writer.sheets["SVIs"], "SVIs")
        df_int.to_excel(writer, sheet_name="Ints", startrow=0)
        formatting(writer.sheets["Ints"], "Ints")
        df_po.to_excel(writer, sheet_name="Po", startrow=0)
        formatting(writer.sheets["Po"], "Po")
        df_route.to_excel(writer, sheet_name="Statics", startrow=0)
        formatting(writer.sheets["Statics"], "Statics")
        df_ipacl.to_excel(writer, sheet_name="IP_ACLs", startrow=0)
        formatting(writer.sheets["IP_ACLs"], "IP_ACLs")
    print("The file was generated:  {}".format(wbname))
# ----------------------------------------------------


if __name__ == "__main__":
    print("")
    ##########################################################################
    import argparse
    desc = "This example script generates an Excel file with the information gathered from running-config file from Cisco NXOS."
    separator = "-" * 110

    def indent_formatter(prog): return argparse.RawTextHelpFormatter(
        prog, max_help_position=50)
    parser = argparse.ArgumentParser(
        prog="parse_nxos_conf",
        description=desc,
        formatter_class=indent_formatter,
        epilog=separator)
    parser.add_argument(
        '-f',
        '--folder',
        action='store_true',
        help='Read all config files from /Configs')
    parser.add_argument(
        '-m',
        '--mode',
        action='store',
        help='Optional argument: mode nxos',
        metavar='Mode')
    args = parser.parse_args()
    ##########################################################################
    mode = "nxos"
    if args.mode == "nxos":
        mode = args.mode
    elif args.mode == "ios":
        mode = args.mode
    if args.folder != True:
        root = tk.Tk()
        root.withdraw()
        path = filedialog.askopenfilename()
        # NXOS running config file
        try:
            parse = CiscoConfParse(path, syntax=mode)
        except UnicodeDecodeError as uni:
            print("USE A TEXT FILE!!! Closing...")
            exit()
        wbname = path.split("/")[-1]
        wbname = wbname[:wbname.rfind(".")] + ".xlsx"
        print("Working on directory: ", os.path.dirname(path))
        print("Using the folowing files: ", os.path.basename(path))
        print("Going to write in: ", os.path.dirname(path))
        conf_excel(os.path.join(os.path.dirname(path), wbname), mode)
    else:
        path = os.path.join(os.getcwd(), "Configs")
        onlyfiles = [f for f in listdir(path) if isfile(
            join(path, f)) and not f.startswith(".")]
        print("Working on directory: ", path)
        print("Using the folowing files: ", onlyfiles)
        path2 = os.path.join(os.getcwd(), "Outputs")
        if not os.path.isdir(path2):
            os.mkdir(path2)
        print("Going to write in: ", path2)
        for i in onlyfiles:
            wbname = i[:i.rfind(".")] + ".xlsx"
            try:
                parse = CiscoConfParse(os.path.join(path, i), syntax=mode)
            except UnicodeDecodeError as uni:
                print("USE A TEXT FILE!!! Closing...")
                exit()
            conf_excel(os.path.join(path2, wbname), mode)
    print("")
