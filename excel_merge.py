import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import tkinter as tk
from tkinter import filedialog

#----------------------------------------------------
def formatting (ws, sheet):
	i = 0
	cmax = [6] * ws.max_column
	max_column = cmax
	thin = Side(border_style = "thin", color = "000000")
	for r in ws["A1:"+chr(ws.max_column+64)+str(ws.max_row)]:
		for c in r:
			c.border = Border(bottom = thin, top = thin, right = thin, left = thin)
			if c.value == "X":
				c.fill = PatternFill(start_color = "FB2C57", end_color = "FB2C57", fill_type = "solid")
#Ajuste del ancho de columnas
			try:
				lenght = len(c.value)
			except TypeError as uni:
				lenght = 0
			cmax[i] = max(lenght, cmax[i])
			i += 1
		i = 0
		for j in range(ws.max_column):
			max_column[j] = max(max_column[j], cmax[j]) 
	for j in range(len(max_column)):
		ws.column_dimensions[chr(j+65)].width = max_column[j]
#----------------------------------------------------

sheet = "Hoja"
index = "Name"
sort = index

root = tk.Tk()
root.withdraw()
path1 = filedialog.askopenfilename()
path2 = filedialog.askopenfilename()
xls1 = pd.ExcelFile(path1)
xls2 = pd.ExcelFile(path2)
wb = openpyxl.Workbook()
ws=wb.active
with pd.ExcelWriter("excel_merge.xlsx", engine="openpyxl") as writer:
	writer.book=wb
	df_xls1 = pd.read_excel(xls1, sheet).fillna("#-")
	df_xls2 = pd.read_excel(xls2, sheet).fillna("#-")
	df_xls1 = df_xls1.set_index(index)
	df_xls2 = df_xls2.set_index(index)
	print(df_xls1)
	print(df_xls2)
	df = pd.concat([df_xls1, df_xls2], axis=1).fillna("X")
	df = df.replace(to_replace="#-", value=np.nan)
	df = df.sort_values(by=sort, ascending=True)
	df.to_excel(writer, sheet_name=sheet, startrow = 1)
	formatting(writer.sheets[sheet], sheet)




